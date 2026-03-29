import customtkinter as ctk
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import openpyxl
import threading

# UI Theme Setup
ctk.set_appearance_mode("System")  # Modes: "System", "Dark", "Light"
ctk.set_default_color_theme("blue")

class WhatsAppAutomationUI(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("WhatsApp Automation Tool")
        self.geometry("500x550")

        # --- UI Elements ---
        self.label_title = ctk.CTkLabel(self, text="WhatsApp Bulk Messenger", font=("Roboto", 24))
        self.label_title.pack(pady=20)

        # File Selection
        self.file_path = ""
        self.btn_browse = ctk.CTkButton(self, text="Select Excel File", command=self.browse_file)
        self.btn_browse.pack(pady=10)
        
        self.label_file = ctk.CTkLabel(self, text="No file selected", font=("Roboto", 10))
        self.label_file.pack()

        # Row and Column Inputs
        self.entry_row = ctk.CTkEntry(self, placeholder_text="Start Row Number (e.g. 2)")
        self.entry_row.pack(pady=10, padx=20, fill="x")

        self.entry_col = ctk.CTkEntry(self, placeholder_text="Phone Column Number (e.g. 1)")
        self.entry_col.pack(pady=10, padx=20, fill="x")

        # Message Input
        self.text_msg = ctk.CTkTextbox(self, height=100)
        self.text_msg.pack(pady=10, padx=20, fill="x")
        self.text_msg.insert("0.0", "Type your message here...")

        # Start Button
        self.btn_start = ctk.CTkButton(self, text="Start Sending", command=self.run_automation_thread, fg_color="green", hover_color="#054d05")
        self.btn_start.pack(pady=20)

    def browse_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.file_path:
            self.label_file.configure(text=f"Selected: {self.file_path.split('/')[-1]}")

    def run_automation_thread(self):
        # Threading is used so the UI doesn't freeze while Selenium runs
        t = threading.Thread(target=self.start_automation)
        t.start()

    def start_automation(self):
        # Get data from UI
        s_row = int(self.entry_row.get())
        column = int(self.entry_col.get())
        msg = self.text_msg.get("0.0", "end").strip()

        if not self.file_path or not msg:
            messagebox.showerror("Error", "Please select file and enter message")
            return

        try:
            driver = webdriver.Chrome()
            driver.maximize_window()
            driver.get("https://web.whatsapp.com")
            
            # Wait for Login (Manual)
            messagebox.showinfo("Login", "Please scan QR code, then click OK here.")

            workbook = openpyxl.load_workbook(filename=self.file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=s_row, max_row=sheet.max_row, min_col=column, max_col=column, values_only=True):
                num = row[0]
                if num is None: continue

                link = f"https://web.whatsapp.com/send?phone=91{num}"
                driver.get(link)
                time.sleep(15) # Optimized wait time

                try:
                    # XPath might need updating based on WhatsApp's UI changes
                    input_box = driver.find_element(By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')
                    input_box.send_keys(msg + Keys.ENTER)
                    print(f"Sent to {num}")
                    time.sleep(3)
                except Exception as e:
                    print(f"Could not send to {num}: {e}")

            driver.quit()
            messagebox.showinfo("Success", "All messages sent!")
        except Exception as e:
            messagebox.showerror("Execution Error", str(e))

if __name__ == "__main__":
    app = WhatsAppAutomationUI()
    app.mainloop()